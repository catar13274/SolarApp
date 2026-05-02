"""Materials API endpoints."""

import os
from datetime import datetime
from io import BytesIO
from typing import List, Optional
import unicodedata

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlmodel import Session, select

from ..database import get_tenant_display_name
from ..deps import get_session, resolve_tenant_code
from ..models import Material, Stock

router = APIRouter(prefix="/api/v1/materials", tags=["materials"])
TVA_RATE = 0.21


def _normalize_header(value: str) -> str:
    text = unicodedata.normalize("NFKD", value or "").encode("ascii", "ignore").decode("ascii")
    return " ".join(text.strip().lower().split())


def _firm_label(tenant_key: Optional[str]) -> str:
    if tenant_key:
        return get_tenant_display_name(tenant_key)
    return os.getenv("SOLARAPP_LEGACY_COMPANY_LABEL", "SolarApp")


@router.get("/", response_model=List[dict])
def list_materials(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=100),
    search: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    tenant_key: Optional[str] = Depends(resolve_tenant_code),
    session: Session = Depends(get_session),
):
    """List materials for the current tenant DB (or legacy single DB)."""
    query = select(Material)

    if search:
        query = query.where(
            (Material.name.contains(search))
            | (Material.sku.contains(search))
            | ((Material.description.is_not(None)) & (Material.description.contains(search)))
        )

    if category:
        query = query.where(Material.category == category)

    query = query.order_by(Material.name.asc()).offset(skip).limit(limit)
    materials = session.exec(query).all()

    firm = _firm_label(tenant_key)

    result = []
    for material in materials:
        stock = session.exec(select(Stock).where(Stock.material_id == material.id)).first()

        material_dict = material.model_dump()
        material_dict["current_stock"] = stock.quantity if stock else 0
        material_dict["company_display_name"] = firm
        result.append(material_dict)

    return result


@router.get("/export")
def export_materials(
    tenant_key: Optional[str] = Depends(resolve_tenant_code),
    session: Session = Depends(get_session),
):
    """Export materials to an Excel file."""
    materials = session.exec(select(Material).order_by(Material.name.asc())).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Materials"
    headers = [
        "Cod Produs",
        "Denumire Produs",
        "Cant.",
        "U.M.",
        "Pret Unit. (fara TVA)",
        "Taxa Verde (RON)",
        "Valoare (fara TVA)",
        "TVA (21%)",
        "Valoare cu TVA",
        "Firma (context)",
    ]
    sheet.append(headers)

    firm = _firm_label(tenant_key)

    for material in materials:
        stock = session.exec(select(Stock).where(Stock.material_id == material.id)).first()
        quantity = float(stock.quantity if stock else 0.0)
        unit_price = float(material.unit_price or 0.0)
        green_tax = 0.0
        value_without_vat = (quantity * unit_price) + green_tax
        vat_value = value_without_vat * TVA_RATE
        value_with_vat = value_without_vat + vat_value
        sheet.append(
            [
                material.sku,
                material.name,
                quantity,
                material.unit,
                unit_price,
                green_tax,
                value_without_vat,
                vat_value,
                value_with_vat,
                firm,
            ]
        )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"materials_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/import")
async def import_materials(
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
):
    """Import materials from an Excel file (current tenant / legacy DB)."""
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported for import.")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
        sheet = workbook.active
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {exc}") from exc

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel file does not contain any rows.")

    raw_headers = [_normalize_header(str(cell) if cell is not None else "") for cell in rows[0]]

    header_aliases = {
        "sku": ["cod produs", "sku"],
        "name": ["denumire produs", "name", "nume produs"],
        "quantity": ["cant.", "cant", "cantitate", "qty", "quantity"],
        "unit": ["u.m.", "u.m", "um", "unit", "u m"],
        "unit_price": ["pret unit. (fara tva)", "pret unitar", "unit_price", "pret unit fara tva"],
        "company": ["compania", "company", "firma", "firma (context)"],
        "green_tax": ["taxa verde (ron)", "taxa verde"],
        "value_without_vat": ["valoare (fara tva)"],
        "vat": ["tva (21%)", "tva"],
        "value_with_vat": ["valoare cu tva"],
    }

    header_index = {}
    for canonical, aliases in header_aliases.items():
        idx = next((i for i, col in enumerate(raw_headers) if col in aliases), None)
        if idx is not None:
            header_index[canonical] = idx

    required_headers = {"sku", "name", "quantity", "unit", "unit_price"}
    if not required_headers.issubset(set(header_index.keys())):
        raise HTTPException(
            status_code=400,
            detail=(
                "Missing required columns. Required: "
                "Cod Produs, Denumire Produs, Cant., U.M., Pret Unit. (fara TVA)"
            ),
        )

    created = 0
    updated = 0
    skipped = 0
    errors: List[str] = []

    for row_num, row in enumerate(rows[1:], start=2):
        if row is None:
            skipped += 1
            continue

        def get_val(key: str):
            if key not in header_index:
                return None
            idx = header_index[key]
            return row[idx] if idx < len(row) else None

        sku = str(get_val("sku") or "").strip()
        name = str(get_val("name") or "").strip()
        unit = str(get_val("unit") or "").strip() or "buc"

        if not sku or not name:
            skipped += 1
            continue

        try:
            unit_price = float(get_val("unit_price") or 0)
            quantity = float(get_val("quantity") or 0)
        except (ValueError, TypeError):
            errors.append(f"Row {row_num}: invalid numeric values for Cant./Pret Unit.")
            continue

        existing = session.exec(select(Material).where(Material.sku == sku)).first()
        if existing:
            existing.name = name
            existing.unit = unit
            existing.unit_price = unit_price
            existing.updated_at = datetime.utcnow()
            session.add(existing)
            stock = session.exec(select(Stock).where(Stock.material_id == existing.id)).first()
            if stock:
                stock.quantity = quantity
                session.add(stock)
            else:
                stock = Stock(
                    material_id=existing.id,
                    quantity=quantity,
                    location="Main Warehouse",
                )
                session.add(stock)
            updated += 1
        else:
            material = Material(
                name=name,
                sku=sku,
                description=None,
                category="other",
                unit=unit,
                unit_price=unit_price,
                min_stock=0,
                created_at=datetime.utcnow(),
                updated_at=datetime.utcnow(),
            )
            session.add(material)
            session.flush()

            stock = Stock(material_id=material.id, quantity=quantity, location="Main Warehouse")
            session.add(stock)
            created += 1

    session.commit()
    return {
        "message": "Import completed",
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
    }


@router.get("/{material_id}", response_model=dict)
def get_material(
    material_id: int,
    tenant_key: Optional[str] = Depends(resolve_tenant_code),
    session: Session = Depends(get_session),
):
    """Get material details with current stock."""
    material = session.get(Material, material_id)
    if not material:
        raise HTTPException(status_code=404, detail="Material not found")

    stock = session.exec(select(Stock).where(Stock.material_id == material_id)).first()

    material_dict = material.model_dump()
    material_dict["current_stock"] = stock.quantity if stock else 0
    material_dict["stock_location"] = stock.location if stock else "Main Warehouse"
    material_dict["company_display_name"] = _firm_label(tenant_key)

    return material_dict


@router.post("/", response_model=Material)
def create_material(material: Material, session: Session = Depends(get_session)):
    """Create a new material."""
    existing = session.exec(select(Material).where(Material.sku == material.sku)).first()
    if existing:
        raise HTTPException(status_code=400, detail="SKU already exists")

    material.created_at = datetime.utcnow()
    material.updated_at = datetime.utcnow()

    session.add(material)
    session.commit()
    session.refresh(material)

    stock = Stock(
        material_id=material.id,
        quantity=0.0,
        location="Main Warehouse",
    )
    session.add(stock)
    session.commit()

    return material


@router.put("/{material_id}", response_model=Material)
def update_material(
    material_id: int,
    material_update: Material,
    session: Session = Depends(get_session),
):
    """Update an existing material."""
    material = session.get(Material, material_id)
    if not material:
        raise HTTPException(status_code=404, detail="Material not found")

    if material_update.sku != material.sku:
        existing = session.exec(select(Material).where(Material.sku == material_update.sku)).first()
        if existing:
            raise HTTPException(status_code=400, detail="SKU already exists")

    material.name = material_update.name
    material.sku = material_update.sku
    material.description = material_update.description
    material.category = material_update.category
    material.unit = material_update.unit
    material.unit_price = material_update.unit_price
    material.min_stock = material_update.min_stock
    material.updated_at = datetime.utcnow()

    session.add(material)
    session.commit()
    session.refresh(material)

    return material


@router.delete("/{material_id}")
def delete_material(material_id: int, session: Session = Depends(get_session)):
    """Delete a material."""
    material = session.get(Material, material_id)
    if not material:
        raise HTTPException(status_code=404, detail="Material not found")

    stock = session.exec(select(Stock).where(Stock.material_id == material_id)).first()
    if stock:
        session.delete(stock)

    session.delete(material)
    session.commit()

    return {"message": "Material deleted successfully"}
